import openpyxl as xl

def createWorkbook(papers, fileName):

    wb = xl.Workbook()

    for paper in papers:
        createPaperSheet(wb, paper)

    if fileName[-5:] != ".xlsx":
        fileName += ".xlsx"

    wb.save(fileName)

def createPaperSheet(wb, paper):
    wb.create_sheet(title=paper.name)